from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import shutil
import os
import pdfkit

########输入部分########
#定义字典
textbookDictionary=(
(),
(),
("3-2","花瓣 莲蓬 饱胀 破裂 姿势 仿佛 随风 舞蹈 停止 荷花 清香 圆盘 眼前 本领 飘动"),
("3-3","拼凑 吹拂 赶集 聚拢 形成 横掠 偶尔 沾水 疲倦 闲散 纤细 痕迹 乌黑 活泼 春日 清风 洒落 加入 春光 湖面"),
("3-4",""),
("3-5《守株待兔》","守株 待兔 宋朝 耕田 触动 头颈 释放 其中"),
("3-6《》",""),
("3-7《鹿角和鹿腿》","小鹿 池塘 倒映 欣赏 匀称 别致 配得上 传来 哎呀 狮子 追赶 叹气 痛快 精美 没精打采 机灵 机会"),
("3-8《》",""),
("3-9《古诗三首》","符号 欲塑 灵魂 借问 酒家 何处 牧童 兄弟 独自 异乡 佳节"),
("3-10《纸的发明》","伟大 记录 保存 大约 经验 打捞 阿拉伯 欧洲 社会 造纸术 吸收 原料 满足 朝鮮半岛 日本"),
("3-11《》",""),
("3-12《》",""),
("3-13《》",""),
("3-14《》",""),
("3-15《》",""),
("3-16《宇宙的另一边》","宇宙 流淌 秘密 一栋 房子 楼梯 铃声 乘法 思绪 一篇 飞越 星空 相遇 万物"),
("3-17《我变成了一棵树》","形状 狐狸 弯腰 丁零 巧克力 香肠 继续 抬头 麻烦 担心 失望 背包 面包 花生 牛奶 饭菜 排骨"),
("3-18《童年的水墨画》","水墨画 染色 钓竿 扑腾 破碎 波动 浪花 葫芦 清爽 蘑菇 垂柳 扇动 戏耍 松树 松针"),
("3-19《剃头大师》","表弟 胆小鬼 理发 抢夺 骂人 仇人 差不多 付钱 双倍 虽然 电灯泡 一件 衣服 大师 姑父 欢迎 摆布 过年 央求 天分"),
)

#选择字典
currentCombination=textbookDictionary[10]
#手动输入
#currentCombination=("测试","凯旋 夸大 打开")

########输入部分结束########

#定义函数，将汉字转换为拼音
def translate_to_pinyin():
    from pypinyin import pinyin, lazy_pinyin, Style
    pinyin_text = ' '.join(lazy_pinyin(item,style=Style.TONE))
    return pinyin_text

#定义函数，单元格行列转换为单元格序号，如row1，column1=A1
def get_cell_reference(row, col):
    from openpyxl.utils import get_column_letter
    return f"{get_column_letter(col)}{row}"

# 加载现有的Excel文件
file_dir=os.path.abspath(__file__)
folder_dir = os.path.dirname(file_dir)
workbook = load_workbook(f"{folder_dir}/Template.xlsx")

# 加载Excel模版文件，选择工作表
sheet = workbook['Sheet1']

# 使用 split() 方法将其分割成数组
text_array = currentCombination[1].split()

# 定义一个换行的阈值（例如，5个元素后换行）
line_length = 5

# 初始化计数器
count = 0

# 初始化行索引
row_index = 2

sheet["A1"]=currentCombination[0]
# 遍历数组并添加数据到数据框
for item in text_array:
    # 根据计数器确定要添加数据的列
    column_name = count + 1
    cell_location = get_cell_reference(row_index, column_name)
    img_location = get_cell_reference(row_index+1, column_name)
    answer_location = get_cell_reference(row_index+14, column_name)

    #填入拼音
    sheet[cell_location] = translate_to_pinyin()

    #填入图片
    # 为每次添加创建一个新的图片文件副本
    new_image_path = f'temp_image_{item}.png'
    shutil.copy(f"{folder_dir}/Pic/2zi.png", new_image_path)
    # 创建Image对象
    img = Image(new_image_path)
    # 调整图片大小
    img.width = 116  # 宽度，单位为像素
    img.height = 52 # 高度，单位为像素
    sheet.add_image(img, img_location)

    #填入答案
    sheet[answer_location].value = item
    
    #print(cell_location)
    #print(answer_location)

    #计数
    count += 1
    # 判断是否需要换行
    if count >= line_length:
        count = 0  # 重置计数器
        row_index += 2  # 换到下一行
        

# 保存文件
workbook.save(f"{folder_dir}/{currentCombination[0]}.xlsx")

# 删除临时图片文件
for item in text_array:
    temp_image_path = f'temp_image_{item}.png'
    if os.path.exists(temp_image_path):
        os.remove(temp_image_path)